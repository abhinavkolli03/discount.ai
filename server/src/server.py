import pandas as pd
import numpy as np
from dotenv import load_dotenv
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from flask import Flask, jsonify
from flask_cors import CORS
import boto3
from botocore.exceptions import NoCredentialsError
import os
from datetime import datetime

load_dotenv()

POLYGON_API_KEY = os.getenv('POLYGON_API_KEY')
AWS_ACCESS_KEY_ID = os.getenv('AWS_ACCESS_KEY_ID')
AWS_SECRET_ACCESS_KEY = os.getenv('AWS_SECRET_ACCESS_KEY')
BUCKET_NAME = os.getenv('BUCKET_NAME')

s3 = boto3.client('s3', aws_access_key_id=AWS_ACCESS_KEY_ID, aws_secret_access_key=AWS_SECRET_ACCESS_KEY, region_name='us-east-1')


app = Flask(__name__)
CORS(app, resources={r"/download*": {"origins": "http://localhost:3000"}})

def fetch_financial_data(ticker):
    print("now we are here! oh boy")
    url = f"https://api.polygon.io/vX/reference/financials?ticker={ticker}&timeframe=annual&apiKey={POLYGON_API_KEY}"
    response = requests.get(url)
    if response.status_code == 200:
        print("returned?")
        return response.json()
    else:
        raise Exception(f"Error fetching data: {response.text}")

def extract_financial_statements(financial_data):
    try:
        all_years_data = []
        for year_data in financial_data['results']:
            balance_sheet = year_data['financials'].get('balance_sheet', {})
            income_statement = year_data['financials'].get('income_statement', {})
            cash_flow_statement = year_data['financials'].get('cash_flow_statement', {})
            
            # Assume SG&A is 60% of Operating Expenses
            operating_expenses = income_statement.get('operating_expenses', {}).get('value', 0)
            sg_and_a = operating_expenses * 0.60

            # Assume Depreciation & Amortization is 10% of PPE
            ppe = balance_sheet.get('noncurrent_assets', {}).get('value', 0)
            depreciation_and_amortization = ppe * 0.10

            extracted_data = {
                'start_date': year_data.get('start_date', None),
                'end_date': year_data.get('end_date', None),
                'filing_date': year_data.get('filing_date', None),
                'balance_sheet': {
                    'total_current_assets': balance_sheet.get('current_assets', {}).get('value', None),
                    'total_current_liabilities': balance_sheet.get('current_liabilities', {}).get('value', None),
                    'total_debt': balance_sheet.get('liabilities', {}).get('value', None),
                    'total_equity': balance_sheet.get('equity', {}).get('value', None),
                    'property_plant_equipment_net': ppe
                },
                'income_statement': {
                    'revenue': income_statement.get('revenues', {}).get('value', None),
                    'gross_profit': income_statement.get('gross_profit', {}).get('value', None),
                    'operating_expenses': operating_expenses,
                    'sg_and_a': sg_and_a,
                    'depreciation_and_amortization': depreciation_and_amortization,
                    'income_tax_expense': income_statement.get('income_tax_expense_benefit', {}).get('value', None),
                    'income_before_tax': income_statement.get('income_loss_from_continuing_operations_before_tax', {}).get('value', None),
                    'cost_of_revenue': income_statement.get('cost_of_revenue', {}).get('value', None),
                    'operating_income': income_statement.get('operating_income_loss', {}).get('value', None),
                    'interest_expense': income_statement.get('interest_expense_operating', {}).get('value', None)
                },
                'cash_flow_statement': {
                    'capital_expenditure': cash_flow_statement.get('net_cash_flow_from_investing_activities', {}).get('value', None)
                }
            }
            all_years_data.append(extracted_data)

        return all_years_data

    except Exception as e:
        print(f"An error occurred while extracting financial statements: {e}")
        return None

def file_exists_in_s3(ticker):
    try:
        response = s3.list_objects_v2(Bucket=BUCKET_NAME, Prefix=f"{ticker}/")
        return 'Contents' in response
    except Exception as e:
        print(f"Error checking S3: {e}")
        return False
    
def save_to_csv(ticker, all_years_data):
    print("did we make it here?")
    
    # Lists to store data for each type of financial statement
    income_statements = []
    balance_sheets = []
    cash_flow_statements = []

    for year_data in all_years_data:
        income_statements.append(year_data['income_statement'])
        balance_sheets.append(year_data['balance_sheet'])
        cash_flow_statements.append(year_data['cash_flow_statement'])

    # Creating DataFrames for each type of financial statement
    income_statement_df = pd.DataFrame(income_statements)
    balance_sheet_df = pd.DataFrame(balance_sheets)
    cash_flow_statement_df = pd.DataFrame(cash_flow_statements)

    print("oh were def here!")

    # Creating directory for the ticker if it does not exist
    ticker_dir = f"stocks/{ticker}"
    if not os.path.exists(ticker_dir):
        os.makedirs(ticker_dir)

    # Saving to CSV files
    income_statement_df.to_csv(f"{ticker_dir}/income_statement.csv", index=False)
    balance_sheet_df.to_csv(f"{ticker_dir}/balance_sheet.csv", index=False)
    cash_flow_statement_df.to_csv(f"{ticker_dir}/cash_flow_statement.csv", index=False)
    
    # Assuming s3 is a boto3 S3 client and BUCKET_NAME is defined
    s3.upload_file(f"{ticker_dir}/income_statement.csv", BUCKET_NAME, f"{ticker}/income_statement.csv")
    s3.upload_file(f"{ticker_dir}/balance_sheet.csv", BUCKET_NAME, f"{ticker}/balance_sheet.csv")
    s3.upload_file(f"{ticker_dir}/cash_flow_statement.csv", BUCKET_NAME, f"{ticker}/cash_flow_statement.csv")

def generate_excel(ticker):
    try:
        print("Generating Excel file...")
        projection_time_frame = 5
        risk_free_rate = 0.05
        market_rate = 0.10        
        long_term_growth_rate = 0.02

        if file_exists_in_s3(ticker):
            print(f"Files for {ticker} already exist in S3.")
            return f"{ticker}/{datetime.datetime.now().strftime('%Y%m%d')}-{ticker}-dcf-template.xlsx"

        # Fetch financial data using Polygon.io
        financial_data = fetch_financial_data(ticker)
        print("Fetched financial data:")
        
        if not financial_data['results']:
            raise ValueError("No financial data found for the given ticker.")
        
        financial_statements = extract_financial_statements(financial_data)

        # Save financial data to CSV files
        save_to_csv(ticker, financial_statements)

        # Load the latest financial data into DataFrames
        incomeStatement = pd.DataFrame([year['income_statement'] for year in financial_statements][::-1])
        balanceSheet = pd.DataFrame([year['balance_sheet'] for year in financial_statements][::-1])
        cashFlowState = pd.DataFrame([year['cash_flow_statement'] for year in financial_statements][::-1])
        
        print("Income Statement DataFrame:", incomeStatement)
        print("Balance Sheet DataFrame:", balanceSheet)
        print("Cash Flow Statement DataFrame:", cashFlowState)

        # Load template Excel file
        template_sheet_path = './template.xlsx'
        wb = load_workbook(template_sheet_path)
        ws = wb.active

        # variables
        revenue_growth = incomeStatement["revenue"].pct_change().mean()
        gross_margin = (incomeStatement["gross_profit"] / incomeStatement["revenue"]).mean()
        sga_percent = (incomeStatement["sg_and_a"] / incomeStatement["revenue"]).mean()
        debtamortize = incomeStatement["depreciation_and_amortization"].mean()
        tax_rate = incomeStatement["income_tax_expense"].iloc[-1] / incomeStatement["income_before_tax"].iloc[-1]
        cogs_rev = (incomeStatement["cost_of_revenue"] / incomeStatement["revenue"]).mean()
        opex_rev = (incomeStatement["operating_expenses"] / incomeStatement["revenue"]).mean()
        da_rev = (incomeStatement["depreciation_and_amortization"] / incomeStatement["revenue"]).mean()
        tax_ebit = (incomeStatement["income_tax_expense"] / incomeStatement["operating_income"]).mean()
        cap_rev = (cashFlowState["capital_expenditure"] / incomeStatement["revenue"]).mean()

        # get revenue projections
        current_year = datetime.now().year

        # Initialize revenue list with projections
        revenue = [0 for i in range(projection_time_frame+1)]
        revenue[0] = incomeStatement['revenue'].iloc[-1]
        for i in range(1, len(revenue)):
            revenue[i] = revenue[i-1] * (1+revenue_growth)

        # Adjust years for the projections to start from the current year
        projection_years = [current_year + i for i in range(projection_time_frame + 1)]
        print(f"Revenue Projections ({projection_years}): {revenue}")

        # Recalculate revenue growth using historical data if needed
        historical_revenue = incomeStatement['revenue']
        revenue_growth = historical_revenue.pct_change().mean()
        print("Recalculated Revenue Growth:", revenue_growth)


        # get gross profit projections
        gross_profit = [r*gross_margin for r in revenue]

        # get ebit
        sga = [r * sga_percent for r in revenue]
        da = [debtamortize for i in range(projection_time_frame+1)]
        ebit = [gross_profit[i] - sga[i] - da[i] for i in range(len(gross_profit))]

        # get unlevered fcf
        print("Columns in balanceSheet DataFrame: ", balanceSheet.columns)
        print("First few rows of balanceSheet: ", balanceSheet.head())

        balanceSheet.columns = balanceSheet.columns.str.strip()

        try:
            asset_growth = (balanceSheet["total_current_assets"].pct_change()).mean()
            liability_growth = (balanceSheet["total_current_liabilities"].pct_change()).mean()
        except KeyError as e:
            print(f"Error: {e}. Column not found.")
            return

        ppe_growth = balanceSheet['property_plant_equipment_net'].pct_change().mean()

        assets = [0 for i in range(projection_time_frame+1)]
        assets[0] = balanceSheet['total_current_assets'].iloc[-1]
        for i in range(1, len(assets)):
            assets[i] = assets[i-1] * (1+asset_growth)

        liabilities = [0 for i in range(projection_time_frame+1)]
        liabilities[0] = balanceSheet['total_current_liabilities'].iloc[-1]
        for i in range(1, len(liabilities)):
            liabilities[i] = liabilities[i-1] * (1+liability_growth)

        nwc = [assets[i]-liabilities[i] for i in range(len(assets))]
        delta_nwc = [nwc[i]-nwc[i-1] for i in range(1, len(nwc))]
        delta_avg = sum(delta_nwc)/len(delta_nwc)/1000000

        ppe = [0 for i in range(projection_time_frame+1)]
        ppe[0] = balanceSheet['property_plant_equipment_net'].iloc[-1]
        for i in range(1, len(ppe)):
            ppe[i] = liabilities[i-1] * (1+ppe_growth)
        delta_ppe = [nwc[i]-nwc[i-1] for i in range(1, len(ppe))]
        da = da[1:]
        capex = [delta_ppe[i]+da[i] for i in range(len(da))]
        nopat = [e*(1-tax_rate) for e in ebit[1:]]

        unleveredfcf = [nopat[i] + da[i] - capex[i] - delta_nwc[i] for i in range(len(nopat))]
        print("Unlevered FCF:", unleveredfcf)

        # get beta
        betaEndpoint = f"https://api.newtonanalytics.com/stock-beta/?ticker={ticker.lower()}&index=^GSPC&interval=1mo​&observations=10"
        response = requests.get(betaEndpoint)
        if response.status_code == 200:
            betaJson = response.json()
            beta = betaJson["data"]
            print("Beta:", beta)
        else:
            print(f"Error {response.status_code}: {response.text}")
            
        # calculate WACC
        try:
            last_totaldebt_value = balanceSheet['total_debt'].iloc[-1]
            last_totalequity_value = balanceSheet['total_equity'].iloc[-1]
            weightDebt = last_totaldebt_value / (last_totaldebt_value + last_totalequity_value)
            weightEquity = last_totalequity_value / (last_totaldebt_value + last_totalequity_value)

            interestExpense = incomeStatement['interest_expense'].iloc[-1]
            average_total_debt = balanceSheet['total_debt'].mean()
            interest_rate = interestExpense / average_total_debt
            costEquity = risk_free_rate + beta * market_rate

            wacc = weightDebt * interest_rate * (1 - tax_rate) + weightEquity * costEquity
            print("WACC:", wacc)
        except Exception as e:
            print(f"Error in WACC calculation: {e}")

        # get discounted cash flow and equity value
        # THIS EQUITY VALUE MIGHT BE WRONG
        try:
            discountedfcf = [unleveredfcf[i]/(1+wacc)**(i+1) for i in range(len(unleveredfcf))]

            projection_period_value = sum(discountedfcf)
            terminalfcf = unleveredfcf[-1]*(1+wacc)
            print(terminalfcf)
            terminalvalue = terminalfcf/(wacc-long_term_growth_rate)
            companyvalue = terminalvalue + projection_period_value
            print(companyvalue)
            equityvalue = companyvalue - last_totaldebt_value + (balanceSheet["total_current_assets"].iloc[-1] * 1) # Changed 'cashAndCashEquivalents' to 'total_current_assets'
            print("Equity Value: $" + str(round(equityvalue/1000000000, 5)) + " Billion")
        except Exception as e:
            print(f"Error in equity value calculation: {e}")
        
        #might not need this
        # try:
        #     request = requests.get(f"https://api.sec-api.io/float?ticker={ticker.upper()}&token=37a574d69a6bf9a82714b225b153d1a8d3a4cbcb1309")
        #     sharesoutstanding = request.json()["results"]["shares_outstanding"]
        #     print("Implied Share Price: $" + str(round(equityvalue / sharesoutstanding, 2)))
        # except Exception as e:
        #     print(f"Error in retrieving shares outstanding: {e}")


        # Correct the historical_data DataFrame creation
        historical_data = pd.DataFrame({
            'Revenue': incomeStatement['revenue'] / 1000000,
            'Revenue Growth': incomeStatement['revenue'].pct_change().fillna(0),  # Ensure there's no NaN value
            'COGS': incomeStatement['cost_of_revenue'] / 1000000 if "cost_of_revenue" in incomeStatement.columns else 0,
            '% Revenue': (incomeStatement['cost_of_revenue'] / incomeStatement['revenue']).fillna(0) if "cost_of_revenue" in incomeStatement.columns else 0,
            'Gross Profit': incomeStatement['gross_profit'] / 1000000,
            'Operating Costs': incomeStatement['operating_expenses'] / 1000000,
            '% Revenue OpEx': (incomeStatement['operating_expenses'] / incomeStatement['revenue']).fillna(0),
            'EBITDA': incomeStatement['ebitda'] / 1000000 if "ebitda" in incomeStatement.columns else 0,
            'Depreciation and Amortization': incomeStatement['depreciation_and_amortization'] / 1000000 if "depreciation_and_amortization" in incomeStatement.columns else 0,
            '% Revenue D&A': (incomeStatement['depreciation_and_amortization'] / incomeStatement['revenue']).fillna(0) if "depreciation_and_amortization" in incomeStatement.columns else 0,
            "EBIT": incomeStatement['operating_income'] / 1000000,
            'Taxes': incomeStatement['income_tax_expense'] / 1000000,
            '% EBIT': (incomeStatement['income_tax_expense'] / incomeStatement['operating_income']).fillna(0),
            'NOPAT': (incomeStatement['operating_income'] - incomeStatement['income_tax_expense']) / 1000000,
            'ADD: D&A': incomeStatement['depreciation_and_amortization'] / 1000000 if "depreciation_and_amortization" in incomeStatement.columns else 0,
            'CAPEX': cashFlowState['capital_expenditure'] / 1000000 if "capital_expenditure" in cashFlowState.columns else 0,
            '% Revenue CapEx': (cashFlowState['capital_expenditure'] / incomeStatement['revenue'] * -1).fillna(0) if "capital_expenditure" in cashFlowState.columns else 0,
            'Change in NWC': pd.Series(delta_nwc).reindex(range(len(incomeStatement))).fillna(0) / 1000000,
            'Unlevered FCF': pd.Series(unleveredfcf).reindex(range(len(incomeStatement))).fillna(0) / 1000000,
        })



        transposed_data = historical_data.transpose()

        wb = load_workbook(template_sheet_path)
        ws = wb.active  # or ws = wb['SheetName']

        starting_row = 4
        starting_col = 4

        # Iterate over DataFrame rows (which were originally columns)
        for i, (index, row) in enumerate(transposed_data.iterrows()):
            row_num = starting_row + i

            # Iterate over DataFrame columns (which were originally rows)
            for j, col_name in enumerate(transposed_data.columns):
                col_num = starting_col + j
                col_letter = get_column_letter(col_num)

                value = row[col_name]
                ws[f'{col_letter}{row_num}'] = value

        for i in range(5):
            col_num = 9 + i
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}{5}'] = revenue_growth
            ws[f'{col_letter}{7}'] = cogs_rev
            ws[f'{col_letter}{10}'] = opex_rev
            ws[f'{col_letter}{13}'] = da_rev
            ws[f'{col_letter}{16}'] = tax_rate
            ws[f'{col_letter}{20}'] = cap_rev
            ws[f'{col_letter}{21}'] = delta_avg

        # Save the workbook
        file_path = f"dcf_final_{ticker}.xlsx"
        wb.save(f"stocks/{ticker}/{file_path}")
        print("Data written to Excel in horizontal format.")
        return file_path
    except Exception as e:
        print(f"An error occurred: {e}")
        raise


@app.route('/')
def home():
    return 'Welcome to the financial analysis tool! Go to /download to download the Excel file.'

@app.route('/download/<ticker>', methods=['GET'])
def download(ticker):
    try:
        file_key = generate_excel(ticker)
        ticker_dir = f"stocks/{ticker}"
        s3.upload_file(f"{ticker_dir}/{file_key}", BUCKET_NAME, f"{ticker}/{file_key}")
        print(file_key)

        # Generate a presigned URL for the uploaded S3 object
        response = s3.generate_presigned_url('get_object',
                                             Params={'Bucket': BUCKET_NAME,
                                                     'Key': f"{ticker}/{file_key}"},
                                             ExpiresIn=3600)
        return jsonify(url=response), 200

    except NoCredentialsError:
        return jsonify(error="Credentials not available"), 400
    except Exception as e:
        return jsonify(error=str(e)), 500

if __name__ == "__main__":
    app.run(debug=True, port=5000)
