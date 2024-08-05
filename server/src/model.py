import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from flask import Flask, send_file, jsonify
from flask_cors import CORS
import boto3
from botocore.exceptions import NoCredentialsError
import os

app = Flask(__name__)
CORS(app, resources={r"/download*": {"origins": "http://localhost:3000"}})

def generate_excel(ticker):
    projection_time_frame = 5
    risk_free_rate = 0.05
    market_rate = 0.10
    long_term_growth_rate = 0.02

    base_path = f'./src/{ticker}'
    absolute_base_path = f'./src'
    income_statement_path = os.path.join(base_path, 'income_statement.csv')
    cash_flow_path = os.path.join(base_path, 'cash_flow.csv')
    balance_sheet_path = os.path.join(base_path, 'balance_sheet.csv')

    # Debug: Print paths
    print(f"Income Statement Path: {income_statement_path}")
    print(f"Cash Flow Path: {cash_flow_path}")
    print(f"Balance Sheet Path: {balance_sheet_path}")

    # Check if files exist
    if not os.path.exists(income_statement_path):
        print(f"Error: {income_statement_path} does not exist.")
        return
    if not os.path.exists(cash_flow_path):
        print(f"Error: {cash_flow_path} does not exist.")
        return
    if not os.path.exists(balance_sheet_path):
        print(f"Error: {balance_sheet_path} does not exist.")
        return

    # Read the CSV files
    incomeStatement = pd.read_csv(income_statement_path).iloc[::-1]
    cashFlowState = pd.read_csv(cash_flow_path).iloc[::-1]
    balanceSheet = pd.read_csv(balance_sheet_path).iloc[::-1]

    template_sheet_path = os.path.join(absolute_base_path, 'template.xlsx')

    # variables
    revenue_growth = incomeStatement["revenue"].pct_change().mean()
    gross_margin = (incomeStatement["grossProfit"] / incomeStatement["revenue"]).mean()
    sga_percent = (incomeStatement["sellingGeneralAndAdministrativeExpenses"] / incomeStatement["revenue"]).mean()
    debtamortize = incomeStatement["depreciationAndAmortization"].mean()
    tax_rate = incomeStatement["incomeTaxExpense"].iloc[-1] / incomeStatement["incomeBeforeTax"].iloc[-1]
    cogs_rev = (incomeStatement["costOfRevenue"] / incomeStatement["revenue"]).mean()
    opex_rev = (incomeStatement["operatingExpenses"] / incomeStatement["revenue"]).mean()
    da_rev = (incomeStatement["depreciationAndAmortization"] / incomeStatement["revenue"]).mean()
    tax_ebit = (incomeStatement["incomeTaxExpense"] / incomeStatement["operatingIncome"]).mean()
    cap_rev = (cashFlowState["capitalExpenditure"] / incomeStatement["revenue"]).mean()

    # get rev projections
    revenue = [0 for i in range(projection_time_frame+1)]
    revenue[0] = incomeStatement['revenue'].iloc[-1]
    for i in range(1, len(revenue)):
        revenue[i] = revenue[i-1] * (1+revenue_growth)

    # get gross profit projections
    gross_profit = [r*gross_margin for r in revenue]

    # get ebit
    sga = [r*sga_percent for r in revenue]
    da = [debtamortize for i in range(projection_time_frame+1)]
    ebit = [gross_profit[i] - sga[i] - da[i] for i in range(len(gross_profit))]

    # get unlevered fcf
    print("Columns in balanceSheet DataFrame: ", balanceSheet.columns)
    print("First few rows of balanceSheet: ", balanceSheet.head())

    balanceSheet.columns = balanceSheet.columns.str.strip()

    try:
        asset_growth = (balanceSheet["totalCurrentAssets"].pct_change()).mean()
        liability_growth = (balanceSheet["totalCurrentLiabilities"].pct_change()).mean()
    except KeyError as e:
        print(f"Error: {e}. Column not found.")
        return

    asset_growth = (balanceSheet["totalCurrentAssets"].pct_change()).mean()
    liability_growth = (balanceSheet["totalCurrentLiabilities"].pct_change()).mean()
    ppe_growth = balanceSheet['propertyPlantEquipmentNet'].pct_change().mean()

    assets = [0 for i in range(projection_time_frame+1)]
    assets[0] = balanceSheet['totalCurrentAssets'].iloc[-1]
    for i in range(1, len(assets)):
        assets[i] = assets[i-1] * (1+asset_growth)

    liabilities = [0 for i in range(projection_time_frame+1)]
    liabilities[0] = balanceSheet['totalCurrentLiabilities'].iloc[-1]
    for i in range(1, len(liabilities)):
        liabilities[i] = liabilities[i-1] * (1+liability_growth)

    nwc = [assets[i]-liabilities[i] for i in range(len(assets))]
    delta_nwc = [nwc[i]-nwc[i-1] for i in range(1, len(nwc))]
    delta_avg = sum(delta_nwc)/len(delta_nwc)/1000000

    ppe = [0 for i in range(projection_time_frame+1)]
    ppe[0] = balanceSheet['propertyPlantEquipmentNet'].iloc[-1]
    for i in range(1, len(ppe)):
        ppe[i] = liabilities[i-1] * (1+ppe_growth)
    delta_ppe = [nwc[i]-nwc[i-1] for i in range(1, len(ppe))]
    da = da[1:]
    capex = [delta_ppe[i]+da[i] for i in range(len(da))]
    nopat = [e*(1-tax_rate) for e in ebit[1:]]

    unleveredfcf = [nopat[i] + da[i] - capex[i] - delta_nwc[i] for i in range(len(nopat))]
    print("Unlevered FCF:", unleveredfcf)

    # get beta
    betaEndpoint = f"https://api.newtonanalytics.com/stock-beta/?ticker={ticker.lower()}&index=^GSPC&interval=1mo​&observations=10"
    response = requests.get(betaEndpoint)
    if response.status_code == 200:
        betaJson = response.json()
        beta = betaJson["data"]
        print("Beta:", beta)
    else:
        print(f"Error {response.status_code}: {response.text}")

    # calculate WACC
    last_totaldebt_value = balanceSheet['totalDebt'].iloc[-1]
    last_totalequity_value = balanceSheet['totalEquity'].iloc[-1]
    weightDebt = last_totaldebt_value / (last_totaldebt_value + last_totalequity_value)
    weightEquity = last_totalequity_value / (last_totaldebt_value + last_totalequity_value)

    interestExpense = incomeStatement['interestExpense'].iloc[-1]
    average_total_debt = balanceSheet['totalDebt'].mean()
    interest_rate = interestExpense / average_total_debt
    costEquity = risk_free_rate + beta * market_rate

    wacc = weightDebt * interest_rate * (1 - tax_rate) + weightEquity * costEquity
    print("WACC:", wacc)

    # get discounted cash flow and equity value
    discountedfcf = [unleveredfcf[i]/(1+wacc)**(i+1) for i in range(len(unleveredfcf))]

    projection_period_value = sum(discountedfcf)
    terminalfcf = discountedfcf[-1]*(1+wacc)
    terminalvalue = terminalfcf/(wacc-long_term_growth_rate)
    companyvalue = terminalvalue+projection_period_value
    equityvalue = companyvalue-last_totaldebt_value + balanceSheet["cashAndCashEquivalents"].iloc[-1]
    print("Equity Value: $" + str(round(equityvalue/1000000000, 5)) + " Billion")

    request = requests.get(f"https://api.sec-api.io/float?ticker={ticker.upper()}&token=37a574d69a6bf9a82714b225b153d1a8d3a4cbcb1309d930f10799c2948f8684")

    sharesoutstanding = request.json()["data"][0]["float"]["outstandingShares"][0]['value']
    print("Implied Share Price: $" + str(round(equityvalue/sharesoutstanding, 2)))

    # Combine these into a single DataFrame
    historical_data = pd.DataFrame({
        'Revenue': incomeStatement['revenue'] / 1000000,
        'Revenue Growth': incomeStatement['revenue'].pct_change(),
        'COGS': incomeStatement['costOfRevenue'] / 1000000,
        '% Revenue': incomeStatement['costOfRevenue'] / incomeStatement['revenue'],
        'Gross Profit': incomeStatement['grossProfit'] / 1000000,
        'Operating Costs': incomeStatement['operatingExpenses'] / 1000000,
        '% Revenue OpEx': incomeStatement['operatingExpenses']/incomeStatement['revenue'],
        'EBITDA': incomeStatement['ebitda'] / 1000000,
        'Depreciation and Amortization': incomeStatement['depreciationAndAmortization'] / 1000000,
        '% Revenue D&A': incomeStatement['depreciationAndAmortization']/incomeStatement['revenue'],
        "EBIT": incomeStatement['operatingIncome'] / 1000000,
        'Taxes': incomeStatement['incomeTaxExpense'] / 1000000,
        '% EBIT': incomeStatement['incomeTaxExpense']/incomeStatement['operatingIncome'],
        'NOPAT': (incomeStatement['operatingIncome'] - incomeStatement['incomeTaxExpense']) / 1000000,
        'ADD: D&A': incomeStatement['depreciationAndAmortization'] / 1000000,
        'CAPEX': cashFlowState['capitalExpenditure'] / 1000000,
        '% Revenue CapEx': cashFlowState['capitalExpenditure']/incomeStatement['revenue'] * -1,
        'Change in NWC': [x / 1000000 for x in delta_nwc],
        'Unlevered FCF': [x / 1000000 for x in unleveredfcf],
    })

    transposed_data = historical_data.transpose()

    wb = load_workbook(template_sheet_path)
    ws = wb.active  # or ws = wb['SheetName']

    starting_row = 4
    starting_col = 4

    # Iterate over DataFrame rows (which were originally columns)
    for i, (index, row) in enumerate(transposed_data.iterrows()):
        row_num = starting_row + i

        # Iterate over DataFrame columns (which were originally rows)
        for j, col_name in enumerate(transposed_data.columns):
            col_num = starting_col + j
            col_letter = get_column_letter(col_num)

            value = row[col_name]
            ws[f'{col_letter}{row_num}'] = value

    for i in range(5):
        col_num = 9 + i
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}{5}'] = revenue_growth
        ws[f'{col_letter}{7}'] = cogs_rev
        ws[f'{col_letter}{10}'] = opex_rev
        ws[f'{col_letter}{13}'] = da_rev
        ws[f'{col_letter}{16}'] = tax_rate
        ws[f'{col_letter}{20}'] = cap_rev
        ws[f'{col_letter}{21}'] = delta_avg

    # Save the workbook
    wb.save("Populated_template.xlsx")
    print("Data written to Excel in horizontal format.")

@app.route('/')
def home():
    return 'Welcome to the financial analysis tool! Go to /download to download the Excel file.'

@app.route('/download<ticker>')
def download(ticker):
    # Generate the Excel file first
    generate_excel(ticker)

    s3 = boto3.client('s3', aws_access_key_id='AKIAX2QZEOD6PZH34RDB',
                      aws_secret_access_key='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', region_name='us-east-1')

    # Upload the file to the S3 bucket
    try:
        s3.upload_file('Populated_template.xlsx',
                       'discount.ai-storage', f"create_{ticker.upper()}_Populated_template.xlsx")
    except NoCredentialsError:
        return jsonify(error="Credentials not available"), 400

    try:
        # Generate a presigned URL for the uploaded S3 object
        response = s3.generate_presigned_url('get_object',
                                             Params={'Bucket': 'discount.ai-storage',
                                                     'Key': f"create_{ticker.upper()}_Populated_template.xlsx"},
                                             ExpiresIn=3600)
    except NoCredentialsError:
        return jsonify(error="Credentials not available"), 400

    # The response contains the presigned URL
    return jsonify(url=response), 200

if __name__ == "__main__":
    app.run(debug=False, port=5000)
